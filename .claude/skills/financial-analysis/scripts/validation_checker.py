"""
Financial Model Validation Checker

Automated validation suite for financial models. Checks for common errors,
inconsistencies, and quality issues with detailed reporting.

Author: Keith Fletcher
Version: 1.0.0
"""

from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
from enum import Enum
import openpyxl
from openpyxl import load_workbook


class ValidationLevel(Enum):
    """Validation severity levels"""
    CRITICAL = "CRITICAL"  # Model is broken
    ERROR = "ERROR"        # Must fix before production
    WARNING = "WARNING"    # Should review
    INFO = "INFO"          # Informational only


@dataclass
class ValidationResult:
    """Result of a single validation check"""
    check_id: str
    description: str
    level: ValidationLevel
    status: str  # PASS, FAIL, N/A
    details: str
    location: Optional[str] = None


class FinancialModelValidator:
    """Comprehensive financial model validation"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=False)
        self.results: List[ValidationResult] = []
    
    def run_all_checks(self) -> List[ValidationResult]:
        """Run complete validation suite"""
        self._check_structure()
        self._check_formulas()
        self._check_assumptions()
        self._check_balance_sheet()
        self._check_data_quality()
        self._check_formatting()
        
        return self.results
    
    def _add_result(
        self, 
        check_id: str, 
        description: str, 
        level: ValidationLevel, 
        status: str, 
        details: str,
        location: Optional[str] = None
    ):
        """Add validation result"""
        self.results.append(ValidationResult(
            check_id=check_id,
            description=description,
            level=level,
            status=status,
            details=details,
            location=location
        ))
    
    def _check_structure(self):
        """Validate model structure"""
        # Check for required sheets
        required_sheets = ['Assumptions', 'Validation']
        for sheet_name in required_sheets:
            if sheet_name in self.wb.sheetnames:
                self._add_result(
                    'STRUCT-001',
                    f'Required sheet "{sheet_name}" exists',
                    ValidationLevel.INFO,
                    'PASS',
                    f'Sheet {sheet_name} found in workbook'
                )
            else:
                self._add_result(
                    'STRUCT-001',
                    f'Required sheet "{sheet_name}" missing',
                    ValidationLevel.ERROR,
                    'FAIL',
                    f'Sheet {sheet_name} not found. This is a required component.',
                    'Workbook structure'
                )
        
        # Check sheet count (models should be organized)
        if len(self.wb.sheetnames) > 15:
            self._add_result(
                'STRUCT-002',
                'Excessive number of sheets',
                ValidationLevel.WARNING,
                'FAIL',
                f'Model has {len(self.wb.sheetnames)} sheets. Consider consolidating for maintainability.',
                'Workbook structure'
            )
        else:
            self._add_result(
                'STRUCT-002',
                'Reasonable number of sheets',
                ValidationLevel.INFO,
                'PASS',
                f'Model has {len(self.wb.sheetnames)} sheets'
            )
    
    def _check_formulas(self):
        """Validate formula quality"""
        error_cells = []
        hardcoded_cells = []
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        # Check for formula errors
                        if isinstance(cell.value, str):
                            error_indicators = ['#DIV/0!', '#N/A', '#VALUE!', '#REF!', '#NAME?', '#NUM!']
                            for error in error_indicators:
                                if error in str(cell.value):
                                    error_cells.append(f'{sheet_name}!{cell.coordinate}')
                            
                            # Check for hardcoded numbers in formulas (bad practice)
                            if cell.value.startswith('='):
                                # Simple check for numbers in formulas
                                import re
                                numbers = re.findall(r'[\d.]+', cell.value)
                                # Filter out cell references like A1, B2
                                suspicious_numbers = [n for n in numbers if len(n) > 2 and '.' in n]
                                if suspicious_numbers and sheet_name not in ['Assumptions', 'Cover']:
                                    hardcoded_cells.append(f'{sheet_name}!{cell.coordinate}')
        
        # Report errors
        if error_cells:
            self._add_result(
                'FORMULA-001',
                'Formula errors detected',
                ValidationLevel.CRITICAL,
                'FAIL',
                f'Found {len(error_cells)} cells with formula errors: {", ".join(error_cells[:5])}{"..." if len(error_cells) > 5 else ""}',
                'Formula validation'
            )
        else:
            self._add_result(
                'FORMULA-001',
                'No formula errors',
                ValidationLevel.INFO,
                'PASS',
                'All formulas calculated without errors'
            )
        
        # Report hardcoded values
        if len(hardcoded_cells) > 10:
            self._add_result(
                'FORMULA-002',
                'Excessive hardcoded values in formulas',
                ValidationLevel.WARNING,
                'FAIL',
                f'Found {len(hardcoded_cells)} cells with hardcoded numbers in formulas. Consider moving to Assumptions sheet.',
                'Formula quality'
            )
        elif len(hardcoded_cells) > 0:
            self._add_result(
                'FORMULA-002',
                'Some hardcoded values in formulas',
                ValidationLevel.INFO,
                'PASS',
                f'Found {len(hardcoded_cells)} cells with hardcoded values - review if appropriate'
            )
        else:
            self._add_result(
                'FORMULA-002',
                'No hardcoded values in formulas',
                ValidationLevel.INFO,
                'PASS',
                'All formulas reference assumptions appropriately'
            )
    
    def _check_assumptions(self):
        """Validate assumptions sheet"""
        if 'Assumptions' not in self.wb.sheetnames:
            return
        
        ws = self.wb['Assumptions']
        
        # Check for assumption documentation
        assumption_rows = 0
        documented_sources = 0
        documented_dates = 0
        
        for row in ws.iter_rows(min_row=2):
            if row[0].value and row[1].value:  # Has category and parameter
                assumption_rows += 1
                
                # Check for source documentation (typically column 4 or 5)
                if len(row) > 4 and row[4].value:
                    documented_sources += 1
                
                # Check for dates (typically column 5 or 6)
                if len(row) > 5 and row[5].value:
                    documented_dates += 1
        
        if assumption_rows == 0:
            self._add_result(
                'ASSUMP-001',
                'No assumptions found',
                ValidationLevel.ERROR,
                'FAIL',
                'Assumptions sheet exists but contains no documented assumptions',
                'Assumptions!A:Z'
            )
        else:
            # Check documentation completeness
            source_pct = documented_sources / assumption_rows if assumption_rows > 0 else 0
            date_pct = documented_dates / assumption_rows if assumption_rows > 0 else 0
            
            if source_pct < 0.5:
                self._add_result(
                    'ASSUMP-002',
                    'Insufficient assumption source documentation',
                    ValidationLevel.WARNING,
                    'FAIL',
                    f'Only {source_pct:.0%} of assumptions have documented sources. Target: 100%',
                    'Assumptions sheet'
                )
            else:
                self._add_result(
                    'ASSUMP-002',
                    'Good assumption documentation',
                    ValidationLevel.INFO,
                    'PASS',
                    f'{source_pct:.0%} of assumptions have documented sources'
                )
            
            if date_pct < 0.5:
                self._add_result(
                    'ASSUMP-003',
                    'Insufficient assumption date documentation',
                    ValidationLevel.WARNING,
                    'FAIL',
                    f'Only {date_pct:.0%} of assumptions have dates. Target: 100%',
                    'Assumptions sheet'
                )
            else:
                self._add_result(
                    'ASSUMP-003',
                    'Good assumption date tracking',
                    ValidationLevel.INFO,
                    'PASS',
                    f'{date_pct:.0%} of assumptions have dates documented'
                )
    
    def _check_balance_sheet(self):
        """Validate balance sheet if present"""
        if 'Balance Sheet' not in self.wb.sheetnames:
            return
        
        ws = self.wb['Balance Sheet']
        
        # Try to find assets, liabilities, and equity totals
        # This is a simplified check - actual implementation would need to be more sophisticated
        assets_found = False
        liabilities_found = False
        equity_found = False
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    value_str = str(cell.value).lower()
                    if 'total assets' in value_str:
                        assets_found = True
                    if 'total liabilities' in value_str:
                        liabilities_found = True
                    if 'total equity' in value_str:
                        equity_found = True
        
        if assets_found and liabilities_found and equity_found:
            self._add_result(
                'BS-001',
                'Balance sheet structure present',
                ValidationLevel.INFO,
                'PASS',
                'Found all three main balance sheet sections'
            )
            
            # Note: Actual balance check would require identifying specific cells
            self._add_result(
                'BS-002',
                'Balance check required',
                ValidationLevel.INFO,
                'N/A',
                'Manual verification required: Assets = Liabilities + Equity'
            )
        else:
            missing = []
            if not assets_found:
                missing.append('Assets')
            if not liabilities_found:
                missing.append('Liabilities')
            if not equity_found:
                missing.append('Equity')
            
            self._add_result(
                'BS-001',
                'Incomplete balance sheet structure',
                ValidationLevel.WARNING,
                'FAIL',
                f'Missing sections: {", ".join(missing)}',
                'Balance Sheet'
            )
    
    def _check_data_quality(self):
        """Check for data quality issues"""
        negative_revenue = []
        
        # Check for negative revenue (common data error)
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and 'revenue' in str(cell.value).lower():
                        # Check adjacent cells for negative values
                        for offset in range(1, 5):
                            if cell.column + offset <= ws.max_column:
                                adjacent_cell = ws.cell(row=cell.row, column=cell.column + offset)
                                if isinstance(adjacent_cell.value, (int, float)) and adjacent_cell.value < 0:
                                    negative_revenue.append(f'{sheet_name}!{adjacent_cell.coordinate}')
        
        if negative_revenue:
            self._add_result(
                'DATA-001',
                'Negative revenue values found',
                ValidationLevel.ERROR,
                'FAIL',
                f'Found {len(negative_revenue)} cells with negative revenue: {", ".join(negative_revenue[:3])}',
                'Data quality'
            )
        else:
            self._add_result(
                'DATA-001',
                'No negative revenue',
                ValidationLevel.INFO,
                'PASS',
                'All revenue values are positive or zero'
            )
    
    def _check_formatting(self):
        """Check for consistent formatting"""
        # Check for consistent currency formatting
        currency_formats = set()
        percentage_formats = set()
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.number_format:
                        if '$' in cell.number_format:
                            currency_formats.add(cell.number_format)
                        if '%' in cell.number_format:
                            percentage_formats.add(cell.number_format)
        
        if len(currency_formats) > 3:
            self._add_result(
                'FORMAT-001',
                'Inconsistent currency formatting',
                ValidationLevel.WARNING,
                'FAIL',
                f'Found {len(currency_formats)} different currency formats. Standardize for consistency.',
                'Formatting'
            )
        else:
            self._add_result(
                'FORMAT-001',
                'Consistent currency formatting',
                ValidationLevel.INFO,
                'PASS',
                'Currency formatting is reasonably consistent'
            )
    
    def generate_report(self) -> Dict:
        """Generate summary report"""
        summary = {
            'total_checks': len(self.results),
            'passed': sum(1 for r in self.results if r.status == 'PASS'),
            'failed': sum(1 for r in self.results if r.status == 'FAIL'),
            'na': sum(1 for r in self.results if r.status == 'N/A'),
            'critical': sum(1 for r in self.results if r.level == ValidationLevel.CRITICAL and r.status == 'FAIL'),
            'errors': sum(1 for r in self.results if r.level == ValidationLevel.ERROR and r.status == 'FAIL'),
            'warnings': sum(1 for r in self.results if r.level == ValidationLevel.WARNING and r.status == 'FAIL'),
        }
        
        summary['pass_rate'] = summary['passed'] / summary['total_checks'] if summary['total_checks'] > 0 else 0
        
        return summary
    
    def print_report(self):
        """Print validation report to console"""
        print(f"\n{'='*80}")
        print(f"FINANCIAL MODEL VALIDATION REPORT")
        print(f"File: {self.file_path}")
        print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*80}\n")
        
        summary = self.generate_report()
        
        print(f"Summary:")
        print(f"  Total Checks: {summary['total_checks']}")
        print(f"  Passed: {summary['passed']} ({summary['pass_rate']:.0%})")
        print(f"  Failed: {summary['failed']}")
        print(f"  N/A: {summary['na']}")
        print(f"\nIssues by Severity:")
        print(f"  Critical: {summary['critical']}")
        print(f"  Errors: {summary['errors']}")
        print(f"  Warnings: {summary['warnings']}")
        
        # Print failures by level
        for level in [ValidationLevel.CRITICAL, ValidationLevel.ERROR, ValidationLevel.WARNING]:
            failures = [r for r in self.results if r.level == level and r.status == 'FAIL']
            if failures:
                print(f"\n{level.value} Issues:")
                for result in failures:
                    print(f"  [{result.check_id}] {result.description}")
                    print(f"    Details: {result.details}")
                    if result.location:
                        print(f"    Location: {result.location}")
        
        print(f"\n{'='*80}")
        
        # Overall assessment
        if summary['critical'] > 0:
            print("ASSESSMENT: Model has CRITICAL issues that must be fixed")
        elif summary['errors'] > 0:
            print("ASSESSMENT: Model has ERRORS that should be fixed before production")
        elif summary['warnings'] > 0:
            print("ASSESSMENT: Model is functional but has WARNINGS to review")
        else:
            print("ASSESSMENT: Model passes all validation checks")
        
        print(f"{'='*80}\n")


def validate_model(file_path: str, print_report: bool = True) -> Dict:
    """
    Validate a financial model.
    
    Args:
        file_path: Path to Excel file to validate
        print_report: Whether to print detailed report
    
    Returns:
        Dictionary with validation summary
    """
    validator = FinancialModelValidator(file_path)
    validator.run_all_checks()
    
    if print_report:
        validator.print_report()
    
    return validator.generate_report()


if __name__ == '__main__':
    # Example usage
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        validate_model(file_path)
    else:
        print("Usage: python validation_checker.py <path_to_model.xlsx>")
