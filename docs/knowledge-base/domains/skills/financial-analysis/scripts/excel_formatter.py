"""
Excel Formatting Utilities

Standardized formatting functions for professional financial models.
Provides consistent styling, color coding, and layout patterns.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ModelFormatter:
    """Professional formatting utilities for Excel financial models."""

    # Standard color palette
    COLORS = {
        "input": "FFFF00",  # Yellow - User inputs
        "calculation": "E0E0E0",  # Light gray - Formulas
        "output": "90EE90",  # Light green - Key results
        "header": "4472C4",  # Blue - Headers
        "validation": "FFB6C1",  # Light pink - Checks
        "hardcoded": "FFA500",  # Orange - Constants
        "error": "FF0000",  # Red - Errors
        "pass": "00FF00",  # Green - Passed checks
        "white": "FFFFFF",  # White
        "black": "000000",  # Black
    }

    # Standard fonts
    FONTS = {
        "title": Font(name="Calibri", size=18, bold=True),
        "heading": Font(name="Calibri", size=14, bold=True),
        "subheading": Font(name="Calibri", size=12, bold=True),
        "header": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "body": Font(name="Calibri", size=11),
        "body_bold": Font(name="Calibri", size=11, bold=True),
    }

    # Standard alignments
    ALIGNMENTS = {
        "left": Alignment(horizontal="left", vertical="center"),
        "center": Alignment(horizontal="center", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
        "wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
    }

    # Standard borders
    BORDERS = {
        "thin": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "medium": Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        ),
        "bottom_double": Border(bottom=Side(style="double")),
        "bottom_thick": Border(bottom=Side(style="thick")),
    }

    @staticmethod
    def format_title(ws, cell: str, title: str):
        """
        Format a title cell.

        Args:
            ws: Worksheet object
            cell: Cell reference (e.g., 'A1')
            title: Title text
        """
        ws[cell] = title
        ws[cell].font = ModelFormatter.FONTS["title"]
        ws[cell].alignment = ModelFormatter.ALIGNMENTS["left"]

    @staticmethod
    def format_heading(ws, cell: str, heading: str):
        """
        Format a heading cell.

        Args:
            ws: Worksheet object
            cell: Cell reference
            heading: Heading text
        """
        ws[cell] = heading
        ws[cell].font = ModelFormatter.FONTS["heading"]
        ws[cell].alignment = ModelFormatter.ALIGNMENTS["left"]

    @staticmethod
    def format_header_row(ws, row: int, headers: list[str], start_col: int = 1):
        """
        Format a header row with blue background and white text.

        Args:
            ws: Worksheet object
            row: Row number
            headers: List of header texts
            start_col: Starting column (default 1 = A)
        """
        for col_offset, header in enumerate(headers):
            col = start_col + col_offset
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = ModelFormatter.FONTS["header"]
            cell.fill = PatternFill(
                start_color=ModelFormatter.COLORS["header"],
                end_color=ModelFormatter.COLORS["header"],
                fill_type="solid",
            )
            cell.alignment = ModelFormatter.ALIGNMENTS["center"]
            cell.border = ModelFormatter.BORDERS["thin"]

    @staticmethod
    def format_input_cell(ws, cell: str, value=None, number_format: str = None):
        """
        Format an input cell with yellow background.

        Args:
            ws: Worksheet object
            cell: Cell reference
            value: Optional value to set
            number_format: Optional number format string
        """
        if value is not None:
            ws[cell] = value

        ws[cell].fill = PatternFill(
            start_color=ModelFormatter.COLORS["input"],
            end_color=ModelFormatter.COLORS["input"],
            fill_type="solid",
        )
        ws[cell].border = ModelFormatter.BORDERS["thin"]

        if number_format:
            ws[cell].number_format = number_format

    @staticmethod
    def format_calculation_cell(ws, cell: str, formula: str = None, number_format: str = None):
        """
        Format a calculation cell with gray background.

        Args:
            ws: Worksheet object
            cell: Cell reference
            formula: Optional formula to set
            number_format: Optional number format string
        """
        if formula is not None:
            ws[cell] = formula

        ws[cell].fill = PatternFill(
            start_color=ModelFormatter.COLORS["calculation"],
            end_color=ModelFormatter.COLORS["calculation"],
            fill_type="solid",
        )
        ws[cell].border = ModelFormatter.BORDERS["thin"]

        if number_format:
            ws[cell].number_format = number_format

    @staticmethod
    def format_output_cell(
        ws, cell: str, formula: str = None, number_format: str = None, bold: bool = True
    ):
        """
        Format an output cell with green background.

        Args:
            ws: Worksheet object
            cell: Cell reference
            formula: Optional formula to set
            number_format: Optional number format string
            bold: Whether to bold the text
        """
        if formula is not None:
            ws[cell] = formula

        ws[cell].fill = PatternFill(
            start_color=ModelFormatter.COLORS["output"],
            end_color=ModelFormatter.COLORS["output"],
            fill_type="solid",
        )
        ws[cell].border = ModelFormatter.BORDERS["thin"]

        if bold:
            ws[cell].font = ModelFormatter.FONTS["body_bold"]

        if number_format:
            ws[cell].number_format = number_format

    @staticmethod
    def format_total_row(
        ws, row: int, start_col: int, end_col: int, bold: bool = True, double_underline: bool = True
    ):
        """
        Format a total row with bold and underline.

        Args:
            ws: Worksheet object
            row: Row number
            start_col: Starting column
            end_col: Ending column
            bold: Whether to bold
            double_underline: Whether to double underline
        """
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)

            if bold:
                cell.font = ModelFormatter.FONTS["body_bold"]

            if double_underline:
                cell.border = ModelFormatter.BORDERS["bottom_double"]

    @staticmethod
    def format_section_header(ws, cell: str, text: str):
        """
        Format a section header (left-aligned, bold, larger font).

        Args:
            ws: Worksheet object
            cell: Cell reference
            text: Header text
        """
        ws[cell] = text
        ws[cell].font = ModelFormatter.FONTS["subheading"]
        ws[cell].alignment = ModelFormatter.ALIGNMENTS["left"]

    @staticmethod
    def apply_currency_format(ws, cell_range: str, decimals: int = 0):
        """
        Apply currency formatting to a range.

        Args:
            ws: Worksheet object
            cell_range: Range string (e.g., 'B5:F20')
            decimals: Number of decimal places
        """
        if decimals == 0:
            fmt = "$#,##0"
        else:
            fmt = f'$#,##0.{"0" * decimals}'

        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = fmt

    @staticmethod
    def apply_percentage_format(ws, cell_range: str, decimals: int = 1):
        """
        Apply percentage formatting to a range.

        Args:
            ws: Worksheet object
            cell_range: Range string
            decimals: Number of decimal places
        """
        fmt = f'0.{"0" * decimals}%'

        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = fmt

    @staticmethod
    def apply_number_format(ws, cell_range: str, decimals: int = 0, use_thousands: bool = True):
        """
        Apply number formatting to a range.

        Args:
            ws: Worksheet object
            cell_range: Range string
            decimals: Number of decimal places
            use_thousands: Whether to use thousands separator
        """
        if use_thousands:
            if decimals == 0:
                fmt = "#,##0"
            else:
                fmt = f'#,##0.{"0" * decimals}'
        elif decimals == 0:
            fmt = "0"
        else:
            fmt = f'0.{"0" * decimals}'

        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = fmt

    @staticmethod
    def apply_date_format(ws, cell_range: str, format_style: str = "yyyy-mm-dd"):
        """
        Apply date formatting to a range.

        Args:
            ws: Worksheet object
            cell_range: Range string
            format_style: Date format string
        """
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = format_style

    @staticmethod
    def set_column_widths(ws, column_widths: dict):
        """
        Set column widths for multiple columns.

        Args:
            ws: Worksheet object
            column_widths: Dict mapping column letters to widths
                          e.g., {'A': 25, 'B': 15, 'C': 15}
        """
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    @staticmethod
    def auto_adjust_column_widths(ws, min_width: int = 8, max_width: int = 50):
        """
        Auto-adjust column widths based on content.

        Args:
            ws: Worksheet object
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(cell_length, max_length)
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def freeze_panes(ws, cell: str = "B4"):
        """
        Freeze panes at specified cell.

        Args:
            ws: Worksheet object
            cell: Cell where freeze occurs (default 'B4' = freeze first row and column)
        """
        ws.freeze_panes = cell

    @staticmethod
    def add_alternating_row_colors(
        ws,
        start_row: int,
        end_row: int,
        start_col: int = 1,
        end_col: int = None,
        color1: str = "FFFFFF",
        color2: str = "F2F2F2",
    ):
        """
        Add alternating row colors for better readability.

        Args:
            ws: Worksheet object
            start_row: First data row
            end_row: Last data row
            start_col: Starting column
            end_col: Ending column (None = max column)
            color1: First color (default white)
            color2: Second color (default light gray)
        """
        if end_col is None:
            end_col = ws.max_column

        for row in range(start_row, end_row + 1):
            color = color1 if (row - start_row) % 2 == 0 else color2

            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    @staticmethod
    def create_validation_cell(
        ws, cell: str, expected_ref: str, actual_ref: str, tolerance: float = 0.01
    ):
        """
        Create a validation check cell with pass/fail indicator.

        Args:
            ws: Worksheet object
            cell: Cell for validation result
            expected_ref: Cell reference for expected value
            actual_ref: Cell reference for actual value
            tolerance: Acceptable difference
        """
        formula = f"=IF(ABS({expected_ref}-{actual_ref})<={tolerance}, " f'"PASS", "FAIL")'
        ws[cell] = formula

        # Conditional formatting would be applied separately
        # This just sets up the formula

    @staticmethod
    def protect_sheet(
        ws,
        password: str | None = None,
        allow_select_locked: bool = True,
        allow_select_unlocked: bool = True,
    ):
        """
        Protect worksheet while allowing selection.

        Args:
            ws: Worksheet object
            password: Optional password
            allow_select_locked: Allow selecting locked cells
            allow_select_unlocked: Allow selecting unlocked cells
        """
        ws.protection.sheet = True
        ws.protection.password = password
        ws.protection.selectLockedCells = allow_select_locked
        ws.protection.selectUnlockedCells = allow_select_unlocked

    @staticmethod
    def unlock_input_cells(ws, input_ranges: list[str]):
        """
        Unlock specific input ranges (for use with sheet protection).

        Args:
            ws: Worksheet object
            input_ranges: List of range strings to unlock (e.g., ['B5:B20', 'D10'])
        """
        for range_str in input_ranges:
            for row in ws[range_str]:
                for cell in row:
                    cell.protection = openpyxl.styles.Protection(locked=False)

    @staticmethod
    def create_styled_table(
        ws,
        start_row: int,
        headers: list[str],
        data: list[list],
        start_col: int = 1,
        apply_alternating_colors: bool = True,
    ):
        """
        Create a professionally styled table.

        Args:
            ws: Worksheet object
            start_row: Starting row number
            headers: List of header texts
            data: List of data rows (list of lists)
            start_col: Starting column
            apply_alternating_colors: Whether to alternate row colors
        """
        # Create header row
        ModelFormatter.format_header_row(ws, start_row, headers, start_col)

        # Add data
        for row_offset, row_data in enumerate(data):
            row = start_row + row_offset + 1
            for col_offset, value in enumerate(row_data):
                col = start_col + col_offset
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = ModelFormatter.BORDERS["thin"]

        # Apply alternating colors if requested
        if apply_alternating_colors:
            end_row = start_row + len(data)
            end_col = start_col + len(headers) - 1
            ModelFormatter.add_alternating_row_colors(
                ws, start_row + 1, end_row, start_col, end_col
            )


# Convenience functions for common formatting patterns


def create_assumptions_sheet(wb: Workbook, sheet_name: str = "Assumptions"):
    """
    Create a properly formatted assumptions sheet.

    Args:
        wb: Workbook object
        sheet_name: Name for assumptions sheet

    Returns:
        Worksheet object
    """
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Title
    ModelFormatter.format_title(ws, "A1", "Model Assumptions")

    # Headers
    headers = ["Category", "Parameter", "Value", "Unit", "Source", "Date", "Sensitivity", "Notes"]
    ModelFormatter.format_header_row(ws, 3, headers)

    # Set column widths
    ModelFormatter.set_column_widths(
        ws,
        {
            "A": 20,  # Category
            "B": 25,  # Parameter
            "C": 15,  # Value
            "D": 12,  # Unit
            "E": 25,  # Source
            "F": 12,  # Date
            "G": 12,  # Sensitivity
            "H": 40,  # Notes
        },
    )

    return ws


def create_validation_sheet(wb: Workbook, sheet_name: str = "Validation"):
    """
    Create a properly formatted validation sheet.

    Args:
        wb: Workbook object
        sheet_name: Name for validation sheet

    Returns:
        Worksheet object
    """
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Title
    ModelFormatter.format_title(ws, "A1", "Model Validation Checks")

    # Headers
    headers = [
        "Check ID",
        "Check Description",
        "Expected",
        "Actual",
        "Status",
        "Tolerance",
        "Notes",
    ]
    ModelFormatter.format_header_row(ws, 3, headers)

    # Set column widths
    ModelFormatter.set_column_widths(
        ws,
        {
            "A": 12,  # Check ID
            "B": 45,  # Description
            "C": 15,  # Expected
            "D": 15,  # Actual
            "E": 10,  # Status
            "F": 10,  # Tolerance
            "G": 35,  # Notes
        },
    )

    return ws


if __name__ == "__main__":
    # Example usage
    wb = Workbook()
    ws = wb.active
    ws.title = "Example"

    # Title
    ModelFormatter.format_title(ws, "A1", "Financial Model Example")

    # Section header
    ModelFormatter.format_section_header(ws, "A3", "Revenue Projections")

    # Headers
    headers = ["Year", "Revenue", "Growth Rate", "COGS", "Gross Profit"]
    ModelFormatter.format_header_row(ws, 5, headers)

    # Input cells
    ModelFormatter.format_input_cell(ws, "B6", 1000000, "$#,##0")
    ModelFormatter.format_input_cell(ws, "C6", 0.15, "0.0%")

    # Calculation cells
    ModelFormatter.format_calculation_cell(ws, "B7", "=B6*(1+C6)", "$#,##0")
    ModelFormatter.format_calculation_cell(ws, "D6", "=B6*0.6", "$#,##0")

    # Output cell
    ModelFormatter.format_output_cell(ws, "E6", "=B6-D6", "$#,##0")

    # Set column widths
    ModelFormatter.set_column_widths(
        ws,
        {
            "A": 15,
            "B": 15,
            "C": 12,
            "D": 15,
            "E": 15,
        },
    )

    # Save example
    wb.save("/mnt/user-data/outputs/formatting_example.xlsx")
    print("Formatting example created: /mnt/user-data/outputs/formatting_example.xlsx")
