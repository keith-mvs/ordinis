"""
Budget vs Actual Analysis Model Generator

Creates variance analysis workbooks comparing budgeted vs actual performance
with automated variance calculations, flags, and commentary sections.
"""

from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class BudgetLine:
    """Single line item in budget."""

    category: str
    line_item: str
    budget_amount: float
    actual_amount: float | None = None
    is_revenue: bool = True  # True for revenue, False for expense


class BudgetVsActualModel:
    """Generator for budget vs actual variance analysis models."""

    COLORS = {
        "input": "FFFF00",
        "calculation": "E0E0E0",
        "output": "90EE90",
        "header": "4472C4",
        "favorable": "90EE90",
        "unfavorable": "FFB6C1",
        "warning": "FFA500",
    }

    def __init__(self, company_name: str, period: str, budget_lines: list[BudgetLine]):
        """
        Initialize budget vs actual model.

        Args:
            company_name: Company name
            period: Time period (e.g., "Q1 2024", "FY 2024")
            budget_lines: List of budget line items
        """
        self.company_name = company_name
        self.period = period
        self.budget_lines = budget_lines
        self.wb = Workbook()

    def create_model(self, output_path: str) -> str:
        """
        Create complete budget vs actual model.

        Args:
            output_path: File path for saved model

        Returns:
            Path to saved model
        """
        # Remove default sheet
        self.wb.remove(self.wb.active)

        # Create sheets
        self._create_cover_sheet()
        self._create_summary_sheet()
        self._create_detail_sheet()
        self._create_variance_analysis()
        self._create_charts()

        # Save workbook
        self.wb.save(output_path)
        return output_path

    def _create_cover_sheet(self):
        """Create cover sheet with model metadata."""
        ws = self.wb.create_sheet("Cover")

        # Title
        ws["A1"] = f"{self.company_name} - Budget vs Actual Analysis"
        ws["A1"].font = Font(size=18, bold=True)

        # Metadata
        ws["A3"] = "Period:"
        ws["A3"].font = Font(bold=True)
        ws["B3"] = self.period

        ws["A4"] = "Created:"
        ws["A4"].font = Font(bold=True)
        ws["B4"] = datetime.now().strftime("%Y-%m-%d")

        ws["A5"] = "Purpose:"
        ws["A5"].font = Font(bold=True)
        ws["B5"] = "Variance analysis comparing budgeted to actual financial performance"

        # Key metrics summary
        ws["A7"] = "Key Metrics Summary"
        ws["A7"].font = Font(size=12, bold=True)

        # These will reference the summary sheet
        ws["A8"] = "Total Budget Revenue:"
        ws["B8"] = "=Summary!B5"
        ws["B8"].number_format = "$#,##0"

        ws["A9"] = "Total Actual Revenue:"
        ws["B9"] = "=Summary!C5"
        ws["B9"].number_format = "$#,##0"

        ws["A10"] = "Revenue Variance:"
        ws["B10"] = "=Summary!D5"
        ws["B10"].number_format = "$#,##0"

        ws["A11"] = "Revenue Variance %:"
        ws["B11"] = "=Summary!E5"
        ws["B11"].number_format = "0.0%"

        ws["A13"] = "Total Budget Expenses:"
        ws["B13"] = "=Summary!B20"
        ws["B13"].number_format = "$#,##0"

        ws["A14"] = "Total Actual Expenses:"
        ws["B14"] = "=Summary!C20"
        ws["B14"].number_format = "$#,##0"

        ws["A15"] = "Expense Variance:"
        ws["B15"] = "=Summary!D20"
        ws["B15"].number_format = "$#,##0"

        ws["A16"] = "Expense Variance %:"
        ws["B16"] = "=Summary!E20"
        ws["B16"].number_format = "0.0%"

        # Format columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_summary_sheet(self):
        """Create summary sheet with high-level variance analysis."""
        ws = self.wb.create_sheet("Summary")

        ws["A1"] = f"Budget vs Actual Summary - {self.period}"
        ws["A1"].font = Font(size=14, bold=True)

        # Headers
        headers = ["Category", "Budget", "Actual", "Variance ($)", "Variance (%)", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLORS["header"],
                end_color=self.COLORS["header"],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")

        # Revenue section
        ws["A4"] = "REVENUE"
        ws["A4"].font = Font(size=11, bold=True)

        row = 5
        revenue_lines = [bl for bl in self.budget_lines if bl.is_revenue]

        for line in revenue_lines:
            ws.cell(row=row, column=1, value=line.line_item)
            ws.cell(row=row, column=2, value=line.budget_amount)

            # Actual cell - yellow for input
            actual_cell = ws.cell(row=row, column=3)
            actual_cell.value = line.actual_amount if line.actual_amount else ""
            actual_cell.fill = PatternFill(
                start_color=self.COLORS["input"], end_color=self.COLORS["input"], fill_type="solid"
            )

            # Variance ($)
            ws.cell(row=row, column=4, value=f"=C{row}-B{row}")

            # Variance (%)
            ws.cell(row=row, column=5, value=f"=IFERROR(D{row}/B{row}, 0)")

            # Status
            ws.cell(
                row=row,
                column=6,
                value=(
                    f'=IF(E{row}>=0.05, "Favorable", '
                    f'IF(E{row}<=-0.05, "Unfavorable", "On Target"))'
                ),
            )

            row += 1

        # Total Revenue
        revenue_end_row = row - 1
        ws.cell(row=row, column=1, value="Total Revenue").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"=SUM(B5:B{revenue_end_row})")
        ws.cell(row=row, column=3, value=f"=SUM(C5:C{revenue_end_row})")
        ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
        ws.cell(row=row, column=5, value=f"=IFERROR(D{row}/B{row}, 0)")
        ws.cell(
            row=row,
            column=6,
            value=(
                f'=IF(E{row}>=0.05, "Favorable", IF(E{row}<=-0.05, "Unfavorable", "On Target"))'
            ),
        )

        # Highlight total row
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=self.COLORS["output"],
                end_color=self.COLORS["output"],
                fill_type="solid",
            )
            ws.cell(row=row, column=col).font = Font(bold=True)

        # Expense section
        row += 2
        ws[f"A{row}"] = "EXPENSES"
        ws[f"A{row}"].font = Font(size=11, bold=True)

        row += 1
        expense_start_row = row
        expense_lines = [bl for bl in self.budget_lines if not bl.is_revenue]

        for line in expense_lines:
            ws.cell(row=row, column=1, value=line.line_item)
            ws.cell(row=row, column=2, value=line.budget_amount)

            # Actual cell - yellow for input
            actual_cell = ws.cell(row=row, column=3)
            actual_cell.value = line.actual_amount if line.actual_amount else ""
            actual_cell.fill = PatternFill(
                start_color=self.COLORS["input"], end_color=self.COLORS["input"], fill_type="solid"
            )

            # Variance ($) - For expenses, under budget is favorable
            ws.cell(row=row, column=4, value=f"=B{row}-C{row}")

            # Variance (%)
            ws.cell(row=row, column=5, value=f"=IFERROR(D{row}/B{row}, 0)")

            # Status - For expenses, positive variance is favorable (spent less)
            ws.cell(
                row=row,
                column=6,
                value=(
                    f'=IF(E{row}>=0.05, "Favorable", '
                    f'IF(E{row}<=-0.05, "Unfavorable", "On Target"))'
                ),
            )

            row += 1

        # Total Expenses
        expense_end_row = row - 1
        ws.cell(row=row, column=1, value="Total Expenses").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"=SUM(B{expense_start_row}:B{expense_end_row})")
        ws.cell(row=row, column=3, value=f"=SUM(C{expense_start_row}:C{expense_end_row})")
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        ws.cell(row=row, column=5, value=f"=IFERROR(D{row}/B{row}, 0)")
        ws.cell(
            row=row,
            column=6,
            value=(
                f'=IF(E{row}>=0.05, "Favorable", IF(E{row}<=-0.05, "Unfavorable", "On Target"))'
            ),
        )

        # Highlight total row
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=self.COLORS["output"],
                end_color=self.COLORS["output"],
                fill_type="solid",
            )
            ws.cell(row=row, column=col).font = Font(bold=True)

        # Net Income
        row += 2
        total_revenue_row = 5 + len(revenue_lines)
        total_expense_row = row - 1

        ws[f"A{row}"] = "Net Income"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        ws[f"B{row}"] = f"=B{total_revenue_row}-B{total_expense_row}"
        ws[f"C{row}"] = f"=C{total_revenue_row}-C{total_expense_row}"
        ws[f"D{row}"] = f"=C{row}-B{row}"
        ws[f"E{row}"] = f"=IFERROR(D{row}/B{row}, 0)"
        ws[f"F{row}"] = (
            f'=IF(E{row}>=0.05, "Favorable", IF(E{row}<=-0.05, "Unfavorable", "On Target"))'
        )

        # Highlight net income row
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=self.COLORS["output"],
                end_color=self.COLORS["output"],
                fill_type="solid",
            )
            ws.cell(row=row, column=col).font = Font(size=12, bold=True)

        # Number formatting
        for row_num in range(5, row + 1):
            ws.cell(row=row_num, column=2).number_format = "$#,##0"
            ws.cell(row=row_num, column=3).number_format = "$#,##0"
            ws.cell(row=row_num, column=4).number_format = "$#,##0"
            ws.cell(row=row_num, column=5).number_format = "0.0%"

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 15

    def _create_detail_sheet(self):
        """Create detailed analysis sheet with commentary."""
        ws = self.wb.create_sheet("Detail")

        ws["A1"] = "Detailed Variance Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = "This sheet provides detailed analysis of significant variances (>5% or >$50K)"

        # Headers
        headers = [
            "Line Item",
            "Budget",
            "Actual",
            "Variance ($)",
            "Variance (%)",
            "Explanation",
            "Action Required",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLORS["header"],
                end_color=self.COLORS["header"],
                fill_type="solid",
            )

        row = 6
        for line in self.budget_lines:
            if line.actual_amount:
                # Calculate variance
                if line.is_revenue:
                    variance = line.actual_amount - line.budget_amount
                else:
                    variance = line.budget_amount - line.actual_amount

                variance_pct = variance / line.budget_amount if line.budget_amount != 0 else 0

                # Only show significant variances
                if abs(variance_pct) >= 0.05 or abs(variance) >= 50000:
                    ws.cell(row=row, column=1, value=line.line_item)
                    ws.cell(row=row, column=2, value=line.budget_amount)
                    ws.cell(row=row, column=3, value=line.actual_amount)
                    ws.cell(row=row, column=4, value=variance)
                    ws.cell(row=row, column=5, value=variance_pct)

                    # Explanation and action cells - leave blank for user input
                    ws.cell(row=row, column=6).fill = PatternFill(
                        start_color=self.COLORS["input"],
                        end_color=self.COLORS["input"],
                        fill_type="solid",
                    )
                    ws.cell(row=row, column=7).fill = PatternFill(
                        start_color=self.COLORS["input"],
                        end_color=self.COLORS["input"],
                        fill_type="solid",
                    )

                    row += 1

        # Format numbers
        for row_num in range(6, row):
            ws.cell(row=row_num, column=2).number_format = "$#,##0"
            ws.cell(row=row_num, column=3).number_format = "$#,##0"
            ws.cell(row=row_num, column=4).number_format = "$#,##0"
            ws.cell(row=row_num, column=5).number_format = "0.0%"

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 40

    def _create_variance_analysis(self):
        """Create variance analysis with flags and trends."""
        ws = self.wb.create_sheet("Variance Analysis")

        ws["A1"] = "Variance Analysis Dashboard"
        ws["A1"].font = Font(size=14, bold=True)

        # Variance threshold analysis
        ws["A3"] = "Variance Threshold Analysis"
        ws["A3"].font = Font(size=12, bold=True)

        ws["A5"] = "Within 5% of Budget:"
        ws["B5"] = '=COUNTIF(Summary!F:F, "On Target")&" line items"'

        ws["A6"] = "Favorable Variances (>5%):"
        ws["B6"] = '=COUNTIF(Summary!F:F, "Favorable")&" line items"'

        ws["A7"] = "Unfavorable Variances (>5%):"
        ws["B7"] = '=COUNTIF(Summary!F:F, "Unfavorable")&" line items"'

        # Top variances
        ws["A9"] = "Items Requiring Attention"
        ws["A9"].font = Font(size=12, bold=True)

        ws["A10"] = "Review Detail sheet for explanation of significant variances (>5% or >$50K)"

        # Rolling forecast section
        ws["A12"] = "Rolling Forecast Impact"
        ws["A12"].font = Font(size=12, bold=True)

        ws["A14"] = "If current trends continue:"
        ws["A15"] = "Annual Revenue Forecast:"
        ws["A16"] = "Annual Expense Forecast:"
        ws["A17"] = "Annual Net Income Forecast:"

        # These would need to be updated based on actual period
        ws["B15"] = "=Summary!C5 * 4  // Assuming quarterly period"
        ws["B16"] = "=Summary!C20 * 4"
        ws["B17"] = "=B15 - B16"

        for row in [15, 16, 17]:
            ws[f"B{row}"].number_format = "$#,##0"

        # Format columns
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

    def _create_charts(self):
        """Create visualization charts."""
        ws = self.wb.create_sheet("Charts")

        ws["A1"] = "Budget vs Actual Visualizations"
        ws["A1"].font = Font(size=14, bold=True)

        # Budget vs Actual comparison chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Budget vs Actual Comparison"
        chart.y_axis.title = "Amount ($)"
        chart.x_axis.title = "Line Items"

        # Reference data from Summary sheet
        # This is a simplified version - would need adjustment based on actual data
        data = Reference(ws, min_col=2, min_row=4, max_row=20, max_col=3)
        cats = Reference(ws, min_col=1, min_row=5, max_row=20)

        # Note: Actual chart implementation would need the Summary sheet data
        # This is a placeholder showing the pattern
        ws["A3"] = "Charts would be generated here based on Summary data"

        # Column widths
        ws.column_dimensions["A"].width = 50


def create_budget_vs_actual_model(
    company_name: str,
    period: str,
    revenue_items: dict[str, float],
    expense_items: dict[str, float],
    actual_revenue: dict[str, float] | None = None,
    actual_expenses: dict[str, float] | None = None,
    output_path: str = None,
) -> str:
    """
    Quick function to create budget vs actual model.

    Args:
        company_name: Company name
        period: Period (e.g., "Q1 2024")
        revenue_items: Dict of {line_item: budget_amount}
        expense_items: Dict of {line_item: budget_amount}
        actual_revenue: Dict of {line_item: actual_amount}
        actual_expenses: Dict of {line_item: actual_amount}
        output_path: Save location

    Returns:
        Path to saved model
    """
    # Create budget lines
    budget_lines = []

    # Add revenue items
    for item, budget in revenue_items.items():
        actual = actual_revenue.get(item) if actual_revenue else None
        budget_lines.append(
            BudgetLine(
                category="Revenue",
                line_item=item,
                budget_amount=budget,
                actual_amount=actual,
                is_revenue=True,
            )
        )

    # Add expense items
    for item, budget in expense_items.items():
        actual = actual_expenses.get(item) if actual_expenses else None
        budget_lines.append(
            BudgetLine(
                category="Expenses",
                line_item=item,
                budget_amount=budget,
                actual_amount=actual,
                is_revenue=False,
            )
        )

    # Create model
    model = BudgetVsActualModel(company_name, period, budget_lines)

    if output_path is None:
        output_path = (
            f'/mnt/user-data/outputs/{company_name.replace(" ", "_")}'
            f'_Budget_vs_Actual_{period.replace(" ", "_")}.xlsx'
        )

    return model.create_model(output_path)


if __name__ == "__main__":
    # Example usage
    revenue_budget = {
        "Product Sales": 5000000,
        "Service Revenue": 2000000,
        "Other Revenue": 500000,
    }

    expense_budget = {
        "Cost of Goods Sold": 3000000,
        "Sales & Marketing": 1500000,
        "Research & Development": 800000,
        "General & Administrative": 600000,
        "Other Expenses": 200000,
    }

    # Simulate some actual results
    actual_rev = {
        "Product Sales": 5200000,  # 4% over budget (favorable)
        "Service Revenue": 1850000,  # 7.5% under budget (unfavorable)
        "Other Revenue": 520000,  # 4% over budget
    }

    actual_exp = {
        "Cost of Goods Sold": 3100000,  # 3.3% over budget (unfavorable)
        "Sales & Marketing": 1450000,  # 3.3% under budget (favorable)
        "Research & Development": 820000,  # 2.5% over budget
        "General & Administrative": 580000,  # 3.3% under budget
        "Other Expenses": 195000,  # 2.5% under budget
    }

    model_path = create_budget_vs_actual_model(
        company_name="Sample Corp",
        period="Q1 2024",
        revenue_items=revenue_budget,
        expense_items=expense_budget,
        actual_revenue=actual_rev,
        actual_expenses=actual_exp,
    )

    print(f"Budget vs Actual model created: {model_path}")
