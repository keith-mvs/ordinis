"""
Financial Model Validation Utilities

Automated validation checks for Excel financial models to ensure:
- No hardcoded values in calculation cells
- All formulas have error handling
- Consistent formatting and color coding
- Balance sheet balances
- No circular references
- Proper audit trail
"""

from dataclasses import dataclass
import re

from openpyxl import load_workbook


@dataclass
class ValidationResult:
    """Result of a single validation check."""

    check_id: str
    check_description: str
    status: str  # PASS, FAIL, WARNING, INFO
    details: str
    affected_cells: list[str] = None

    def __post_init__(self):
        if self.affected_cells is None:
            self.affected_cells = []


class ModelValidator:
    """Comprehensive financial model validation framework."""

    def __init__(self, workbook_path: str):
        """
        Initialize validator with Excel workbook.

        Args:
            workbook_path: Path to Excel file to validate
        """
        self.workbook_path = workbook_path
        self.wb = load_workbook(workbook_path, data_only=False)
        self.results: list[ValidationResult] = []

    def run_all_checks(self) -> list[ValidationResult]:
        """
        Run all validation checks on the model.

        Returns:
            List of validation results
        """
        self.results = []

        # Structural checks
        self.check_required_sheets()
        self.check_sheet_naming()

        # Formula checks
        self.check_hardcoded_values()
        self.check_error_handling()
        self.check_circular_references()

        # Formatting checks
        self.check_color_coding()
        self.check_number_formatting()

        # Audit trail checks
        self.check_assumptions_documented()
        self.check_version_control()

        # Balance checks
        self.check_balance_sheet_balance()

        return self.results

    def check_required_sheets(self) -> ValidationResult:
        """Verify all required sheets are present."""
        required_sheets = {"Cover", "Assumptions", "Validation", "Documentation"}

        actual_sheets = set(self.wb.sheetnames)
        missing_sheets = required_sheets - actual_sheets

        if missing_sheets:
            result = ValidationResult(
                check_id="STR-001",
                check_description="Required sheets present",
                status="FAIL",
                details=f'Missing required sheets: {", ".join(missing_sheets)}',
            )
        else:
            result = ValidationResult(
                check_id="STR-001",
                check_description="Required sheets present",
                status="PASS",
                details="All required sheets present",
            )

        self.results.append(result)
        return result

    def check_sheet_naming(self) -> ValidationResult:
        """Check for proper sheet naming conventions."""
        issues = []

        for sheet_name in self.wb.sheetnames:
            # Check for spaces at start/end
            if sheet_name != sheet_name.strip():
                issues.append(f"{sheet_name}: Has leading/trailing spaces")

            # Check for special characters
            if re.search(r'[<>:"/\\|?*]', sheet_name):
                issues.append(f"{sheet_name}: Contains invalid characters")

            # Check length
            if len(sheet_name) > 31:
                issues.append(f"{sheet_name}: Name too long (>31 chars)")

        if issues:
            result = ValidationResult(
                check_id="STR-002",
                check_description="Sheet naming conventions",
                status="WARNING",
                details="; ".join(issues),
            )
        else:
            result = ValidationResult(
                check_id="STR-002",
                check_description="Sheet naming conventions",
                status="PASS",
                details="All sheet names follow conventions",
            )

        self.results.append(result)
        return result

    def check_hardcoded_values(self, exclude_sheets: list[str] = None) -> ValidationResult:
        """
        Check for hardcoded values in formula cells.

        Args:
            exclude_sheets: Sheets to exclude from check (e.g., Assumptions)
        """
        if exclude_sheets is None:
            exclude_sheets = ["Cover", "Assumptions", "Documentation", "Validation", "Charts"]

        hardcoded_cells = []

        for sheet_name in self.wb.sheetnames:
            if sheet_name in exclude_sheets:
                continue

            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        # Check if cell contains formula
                        if cell.value.startswith("="):
                            # Look for hardcoded numbers in formula (excluding references)
                            formula = cell.value

                            # Remove cell references
                            formula_clean = re.sub(r"[A-Z]+\d+", "", formula)
                            # Remove named ranges (words without numbers)
                            formula_clean = re.sub(r"[A-Za-z_]+", "", formula_clean)
                            # Remove function names
                            formula_clean = re.sub(r"[A-Z]+\(", "", formula_clean)

                            # Check for remaining numbers (potential hardcoded values)
                            numbers = re.findall(r"\d+\.?\d*", formula_clean)

                            # Filter out common acceptable values
                            acceptable = {"0", "1", "12", "365", "100"}
                            suspicious_numbers = [n for n in numbers if n not in acceptable]

                            if suspicious_numbers:
                                cell_ref = f"{sheet_name}!{cell.coordinate}"
                                hardcoded_cells.append(
                                    f"{cell_ref}: {cell.value} (contains {suspicious_numbers})"
                                )

        if hardcoded_cells:
            result = ValidationResult(
                check_id="FML-001",
                check_description="No hardcoded values in formulas",
                status="WARNING",
                details=f"Found {len(hardcoded_cells)} formulas with potential hardcoded values",
                affected_cells=hardcoded_cells,
            )
        else:
            result = ValidationResult(
                check_id="FML-001",
                check_description="No hardcoded values in formulas",
                status="PASS",
                details="No suspicious hardcoded values found in formulas",
            )

        self.results.append(result)
        return result

    def check_error_handling(self) -> ValidationResult:
        """Check that formulas include proper error handling."""
        formulas_without_error_handling = []

        for sheet_name in self.wb.sheetnames:
            if sheet_name in ["Cover", "Documentation"]:
                continue

            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula = cell.value.upper()

                        # Check for division operations without IFERROR
                        if "/" in formula and "IFERROR" not in formula:
                            cell_ref = f"{sheet_name}!{cell.coordinate}"
                            formulas_without_error_handling.append(
                                f"{cell_ref}: Division without IFERROR"
                            )

                        # Check for VLOOKUP/XLOOKUP without error handling
                        if (
                            ("VLOOKUP" in formula or "XLOOKUP" in formula)
                            and "IFERROR" not in formula
                            and "IFNA" not in formula
                        ):
                            cell_ref = f"{sheet_name}!{cell.coordinate}"
                            formulas_without_error_handling.append(
                                f"{cell_ref}: Lookup without error handling"
                            )

        if formulas_without_error_handling:
            result = ValidationResult(
                check_id="FML-002",
                check_description="Formulas have error handling",
                status="WARNING",
                details=f"Found {len(formulas_without_error_handling)} formulas without error handling",
                affected_cells=formulas_without_error_handling[:50],  # Limit to first 50
            )
        else:
            result = ValidationResult(
                check_id="FML-002",
                check_description="Formulas have error handling",
                status="PASS",
                details="All risky formulas have error handling",
            )

        self.results.append(result)
        return result

    def check_circular_references(self) -> ValidationResult:
        """Check for circular references in the model."""
        # Note: This is difficult to check programmatically without Excel's calculation engine
        # This is a simplified check looking for obvious patterns

        potential_circular = []

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        # Check if formula references its own cell
                        if cell.coordinate in cell.value:
                            cell_ref = f"{sheet_name}!{cell.coordinate}"
                            potential_circular.append(f"{cell_ref}: References itself")

        if potential_circular:
            result = ValidationResult(
                check_id="FML-003",
                check_description="No circular references",
                status="FAIL",
                details=f"Found {len(potential_circular)} potential circular references",
                affected_cells=potential_circular,
            )
        else:
            result = ValidationResult(
                check_id="FML-003",
                check_description="No circular references",
                status="PASS",
                details="No obvious circular references detected",
            )

        self.results.append(result)
        return result

    def check_color_coding(self) -> ValidationResult:
        """Check that proper color coding is applied."""
        # Standard colors
        INPUT_COLOR = "FFFFFF00"  # Yellow
        CALC_COLOR = "FFE0E0E0"  # Light gray
        OUTPUT_COLOR = "FF90EE90"  # Light green

        issues = []

        if "Assumptions" in self.wb.sheetnames:
            ws = self.wb["Assumptions"]

            # Check that assumption input cells are yellow
            for row in ws.iter_rows(min_row=4, max_row=30):
                for cell in row:
                    if cell.column == 2 and cell.value is not None:  # Column B typically has values
                        if cell.fill.start_color.rgb != INPUT_COLOR:
                            issues.append(f"Assumptions!{cell.coordinate}: Input cell not yellow")

        if issues:
            result = ValidationResult(
                check_id="FMT-001",
                check_description="Proper color coding applied",
                status="WARNING",
                details=f"Found {len(issues)} cells with incorrect color coding",
                affected_cells=issues[:20],  # Limit display
            )
        else:
            result = ValidationResult(
                check_id="FMT-001",
                check_description="Proper color coding applied",
                status="INFO",
                details="Color coding check passed (limited validation)",
            )

        self.results.append(result)
        return result

    def check_number_formatting(self) -> ValidationResult:
        """Check for consistent number formatting."""
        issues = []

        for sheet_name in self.wb.sheetnames:
            if sheet_name in ["Cover", "Documentation"]:
                continue

            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        # Check if number formatting is applied
                        if cell.number_format == "General":
                            cell_ref = f"{sheet_name}!{cell.coordinate}"
                            issues.append(f"{cell_ref}: Number has no specific format")

        if len(issues) > 20:  # Threshold for warning
            result = ValidationResult(
                check_id="FMT-002",
                check_description="Consistent number formatting",
                status="WARNING",
                details=f"Found {len(issues)} cells without specific number format",
                affected_cells=issues[:20],
            )
        else:
            result = ValidationResult(
                check_id="FMT-002",
                check_description="Consistent number formatting",
                status="PASS",
                details="Number formatting is generally consistent",
            )

        self.results.append(result)
        return result

    def check_assumptions_documented(self) -> ValidationResult:
        """Check that assumptions are properly documented."""
        if "Assumptions" not in self.wb.sheetnames:
            result = ValidationResult(
                check_id="AUD-001",
                check_description="Assumptions documented",
                status="FAIL",
                details="Assumptions sheet not found",
            )
            self.results.append(result)
            return result

        ws = self.wb["Assumptions"]

        # Check for required columns
        required_columns = ["Parameter", "Value", "Source", "Date"]
        found_columns = []

        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and any(req in str(cell_value) for req in required_columns):
                found_columns.append(str(cell_value))

        missing_columns = [col for col in required_columns if col not in " ".join(found_columns)]

        if missing_columns:
            result = ValidationResult(
                check_id="AUD-001",
                check_description="Assumptions documented",
                status="WARNING",
                details=f'Assumptions sheet missing columns: {", ".join(missing_columns)}',
            )
        else:
            result = ValidationResult(
                check_id="AUD-001",
                check_description="Assumptions documented",
                status="PASS",
                details="Assumptions sheet has proper documentation structure",
            )

        self.results.append(result)
        return result

    def check_version_control(self) -> ValidationResult:
        """Check that version control information is present."""
        if "Cover" not in self.wb.sheetnames:
            result = ValidationResult(
                check_id="AUD-002",
                check_description="Version control present",
                status="FAIL",
                details="Cover sheet not found",
            )
            self.results.append(result)
            return result

        ws = self.wb["Cover"]

        # Look for version-related keywords
        version_keywords = ["version", "date", "author", "created"]
        found_keywords = []

        for row in ws.iter_rows(max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_lower = cell.value.lower()
                    for keyword in version_keywords:
                        if keyword in cell_lower:
                            found_keywords.append(keyword)

        if len(set(found_keywords)) >= 3:
            result = ValidationResult(
                check_id="AUD-002",
                check_description="Version control present",
                status="PASS",
                details="Cover sheet contains version control information",
            )
        else:
            result = ValidationResult(
                check_id="AUD-002",
                check_description="Version control present",
                status="WARNING",
                details="Cover sheet may be missing version control details",
            )

        self.results.append(result)
        return result

    def check_balance_sheet_balance(self) -> ValidationResult:
        """Check that balance sheet balances (Assets = Liabilities + Equity)."""
        if "Balance Sheet" not in self.wb.sheetnames:
            result = ValidationResult(
                check_id="BAL-001",
                check_description="Balance sheet balances",
                status="INFO",
                details="Balance sheet not found in model",
            )
            self.results.append(result)
            return result

        ws = self.wb["Balance Sheet"]

        # Look for total assets and total liabilities+equity rows
        assets_row = None
        liab_equity_row = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=50), 1):
            cell_value = str(row[0].value).lower() if row[0].value else ""

            if "total assets" in cell_value:
                assets_row = row_idx
            elif (
                "total liabilities and equity" in cell_value
                or "total liabilities & equity" in cell_value
            ):
                liab_equity_row = row_idx

        if not assets_row or not liab_equity_row:
            result = ValidationResult(
                check_id="BAL-001",
                check_description="Balance sheet balances",
                status="WARNING",
                details="Could not locate total assets or liabilities+equity rows",
            )
        else:
            result = ValidationResult(
                check_id="BAL-001",
                check_description="Balance sheet balances",
                status="INFO",
                details=(
                    f"Balance check rows identified: Assets row {assets_row}, "
                    f"Liab+Equity row {liab_equity_row}. "
                    f"Verify Validation sheet confirms balance."
                ),
            )

        self.results.append(result)
        return result

    def generate_report(self, output_path: str = None) -> str:
        """
        Generate validation report.

        Args:
            output_path: Path for report file

        Returns:
            Report text
        """
        if not self.results:
            self.run_all_checks()

        # Count results by status
        status_counts = {"PASS": 0, "FAIL": 0, "WARNING": 0, "INFO": 0}

        for result in self.results:
            status_counts[result.status] = status_counts.get(result.status, 0) + 1

        # Generate report
        report_lines = [
            "=" * 80,
            "FINANCIAL MODEL VALIDATION REPORT",
            f"Model: {self.workbook_path}",
            f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
            "=" * 80,
            "",
            "SUMMARY",
            "-" * 80,
            f"Total Checks: {len(self.results)}",
            f'Passed: {status_counts["PASS"]}',
            f'Failed: {status_counts["FAIL"]}',
            f'Warnings: {status_counts["WARNING"]}',
            f'Info: {status_counts["INFO"]}',
            "",
            "DETAILED RESULTS",
            "-" * 80,
        ]

        for result in self.results:
            report_lines.extend(
                [
                    "",
                    f"[{result.status}] {result.check_id}: {result.check_description}",
                    f"Details: {result.details}",
                ]
            )

            if result.affected_cells and len(result.affected_cells) <= 10:
                report_lines.append("Affected cells:")
                for cell in result.affected_cells:
                    report_lines.append(f"  - {cell}")
            elif result.affected_cells and len(result.affected_cells) > 10:
                report_lines.append(
                    f"Affected cells: {len(result.affected_cells)} total (showing first 10)"
                )
                for cell in result.affected_cells[:10]:
                    report_lines.append(f"  - {cell}")

        report_lines.extend(
            [
                "",
                "=" * 80,
                "END OF REPORT",
                "=" * 80,
            ]
        )

        report_text = "\n".join(report_lines)

        if output_path:
            with open(output_path, "w") as f:
                f.write(report_text)

        return report_text


def validate_model(workbook_path: str, report_path: str = None) -> tuple[bool, str]:
    """
    Convenience function to validate a financial model.

    Args:
        workbook_path: Path to Excel model
        report_path: Path for validation report

    Returns:
        Tuple of (passed: bool, report: str)
    """
    validator = ModelValidator(workbook_path)
    results = validator.run_all_checks()

    # Check if any critical failures
    critical_failures = [r for r in results if r.status == "FAIL"]
    passed = len(critical_failures) == 0

    report = validator.generate_report(report_path)

    return passed, report


if __name__ == "__main__":
    # Example usage
    from datetime import datetime
    import sys

    if len(sys.argv) > 1:
        model_path = sys.argv[1]
        report_path = model_path.replace(".xlsx", "_validation_report.txt")

        print(f"Validating model: {model_path}")
        passed, report = validate_model(model_path, report_path)

        print(report)

        if passed:
            print("\n✓ Model passed validation")
        else:
            print("\n✗ Model has critical failures")

        print(f"\nDetailed report saved to: {report_path}")
    else:
        print("Usage: python model_validation.py <path_to_excel_model>")
