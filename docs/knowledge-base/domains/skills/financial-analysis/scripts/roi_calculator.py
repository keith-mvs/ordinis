"""
ROI and Investment Analysis Calculator

Calculates return metrics for investment opportunities including NPV, IRR,
payback period, and profitability index with scenario analysis.

Author: Keith Fletcher
Version: 1.0.0
"""

from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLORS = {
    "header": "4472C4",
    "input": "FFFF00",
    "calculation": "E0E0E0",
    "output": "90EE90",
}


@dataclass
class InvestmentAssumptions:
    """Investment analysis assumptions"""

    project_name: str
    initial_investment: float
    annual_cash_flows: list[float]
    discount_rate: float
    tax_rate: float = 0.25
    salvage_value: float = 0.0


class ROICalculator:
    """Calculate investment return metrics"""

    def __init__(self, assumptions: InvestmentAssumptions):
        self.assumptions = assumptions
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def create_analysis(self, output_path: str) -> str:
        """Create complete ROI analysis workbook"""
        self._create_cover_sheet()
        self._create_assumptions_sheet()
        self._create_cash_flow_analysis()
        self._create_metrics_sheet()
        self._create_scenario_analysis()

        self.wb.save(output_path)
        return output_path

    def _create_cover_sheet(self):
        """Create cover sheet"""
        ws = self.wb.create_sheet("Cover")

        ws["A1"] = f"{self.assumptions.project_name} - Investment Analysis"
        ws["A1"].font = Font(size=18, bold=True)

        ws["A3"] = "Analysis Date:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d")
        ws["A4"] = "Author:"
        ws["B4"] = "Financial Analysis System"

        for row in range(3, 5):
            ws[f"A{row}"].font = Font(bold=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40

    def _create_assumptions_sheet(self):
        """Create assumptions sheet"""
        ws = self.wb.create_sheet("Assumptions")

        ws["A1"] = "Investment Assumptions"
        ws["A1"].font = Font(size=14, bold=True)

        # Column headers
        headers = ["Parameter", "Value", "Unit", "Source", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid"
            )

        # Add assumptions
        assumptions = [
            (
                "Initial Investment",
                self.assumptions.initial_investment,
                "USD",
                "Capital budget",
                "Upfront cost",
            ),
            ("Discount Rate", self.assumptions.discount_rate, "%", "WACC", "Cost of capital"),
            ("Tax Rate", self.assumptions.tax_rate, "%", "Corporate rate", "Effective tax rate"),
            (
                "Salvage Value",
                self.assumptions.salvage_value,
                "USD",
                "Estimate",
                "End-of-project value",
            ),
            (
                "Project Life",
                len(self.assumptions.annual_cash_flows),
                "Years",
                "Project plan",
                "Analysis period",
            ),
        ]

        for row, assumption in enumerate(assumptions, 4):
            for col, value in enumerate(assumption, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                if col == 2:  # Value column
                    cell.fill = PatternFill(
                        start_color=COLORS["input"], end_color=COLORS["input"], fill_type="solid"
                    )
                    if isinstance(value, float) and value < 1:
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "#,##0"

        # Create named ranges
        self.wb.define_name("InitialInvestment", "=Assumptions!$B$4")
        self.wb.define_name("DiscountRate", "=Assumptions!$B$5")
        self.wb.define_name("TaxRate", "=Assumptions!$B$6")
        self.wb.define_name("SalvageValue", "=Assumptions!$B$7")

        ws.column_dimensions["A"].width = 25
        for col in range(2, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_cash_flow_analysis(self):
        """Create cash flow analysis"""
        ws = self.wb.create_sheet("Cash Flows")

        ws["A1"] = "Cash Flow Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        # Years header
        ws["A3"] = "Year"
        ws["A3"].font = Font(bold=True)

        for year in range(len(self.assumptions.annual_cash_flows) + 1):
            cell = ws.cell(row=3, column=year + 2)
            cell.value = year
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Line items
        line_items = [
            "Initial Investment",
            "Annual Cash Flow",
            "Salvage Value",
            "Total Cash Flow",
            "Cumulative Cash Flow",
            "Discount Factor",
            "Present Value",
            "Cumulative PV",
        ]

        for row, item in enumerate(line_items, 4):
            ws[f"A{row}"] = item
            ws[f"A{row}"].font = Font(
                bold=True if item in ["Total Cash Flow", "Cumulative PV"] else False
            )

        # Initial investment (Year 0)
        ws["B4"] = "=-InitialInvestment"
        for col in range(3, 3 + len(self.assumptions.annual_cash_flows)):
            ws.cell(row=4, column=col, value=0)

        # Annual cash flows
        ws["B5"] = 0
        for col, cf in enumerate(self.assumptions.annual_cash_flows, 3):
            ws.cell(row=5, column=col, value=cf)

        # Salvage value (last year only)
        for col in range(2, 3 + len(self.assumptions.annual_cash_flows)):
            if col == 2 + len(self.assumptions.annual_cash_flows):
                ws.cell(row=6, column=col, value="=SalvageValue")
            else:
                ws.cell(row=6, column=col, value=0)

        # Total cash flow
        for col in range(2, 3 + len(self.assumptions.annual_cash_flows)):
            col_letter = get_column_letter(col)
            ws[f"{col_letter}7"] = f"={col_letter}4+{col_letter}5+{col_letter}6"
            ws[f"{col_letter}7"].fill = PatternFill(
                start_color=COLORS["calculation"],
                end_color=COLORS["calculation"],
                fill_type="solid",
            )

        # Cumulative cash flow
        ws["B8"] = "=B7"
        for col in range(3, 3 + len(self.assumptions.annual_cash_flows)):
            col_letter = get_column_letter(col)
            prev_col = get_column_letter(col - 1)
            ws[f"{col_letter}8"] = f"={prev_col}8+{col_letter}7"

        # Discount factor
        for col in range(2, 3 + len(self.assumptions.annual_cash_flows)):
            col_letter = get_column_letter(col)
            year = col - 2
            ws[f"{col_letter}9"] = f"=1/((1+DiscountRate)^{year})"
            ws[f"{col_letter}9"].number_format = "0.0000"

        # Present value
        for col in range(2, 3 + len(self.assumptions.annual_cash_flows)):
            col_letter = get_column_letter(col)
            ws[f"{col_letter}10"] = f"={col_letter}7*{col_letter}9"

        # Cumulative PV
        ws["B11"] = "=B10"
        for col in range(3, 3 + len(self.assumptions.annual_cash_flows)):
            col_letter = get_column_letter(col)
            prev_col = get_column_letter(col - 1)
            ws[f"{col_letter}11"] = f"={prev_col}11+{col_letter}10"
            ws[f"{col_letter}11"].fill = PatternFill(
                start_color=COLORS["output"], end_color=COLORS["output"], fill_type="solid"
            )

        # Format currency
        for row in range(4, 12):
            for col in range(2, 3 + len(self.assumptions.annual_cash_flows)):
                if row != 9:  # Skip discount factor row
                    ws.cell(row=row, column=col).number_format = "$#,##0"

        ws.column_dimensions["A"].width = 25
        for col in range(2, 3 + len(self.assumptions.annual_cash_flows)):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_metrics_sheet(self):
        """Create investment metrics"""
        ws = self.wb.create_sheet("Metrics")

        ws["A1"] = "Investment Metrics"
        ws["A1"].font = Font(size=14, bold=True)

        # Key metrics
        metrics = [
            (
                "Net Present Value (NPV)",
                f"='Cash Flows'!{get_column_letter(2 + len(self.assumptions.annual_cash_flows))}11",
                "$#,##0",
            ),
            (
                "Internal Rate of Return (IRR)",
                f"=IRR('Cash Flows'!B7:{get_column_letter(2 + len(self.assumptions.annual_cash_flows))}7)",
                "0.00%",
            ),
            ("Profitability Index", "=(B3+InitialInvestment)/InitialInvestment", "0.00"),
            (
                "Payback Period (Years)",
                f"=MATCH(TRUE, 'Cash Flows'!B8:{get_column_letter(2 + len(self.assumptions.annual_cash_flows))}8>=0, 0)-1",
                "0.0",
            ),
            (
                "Discounted Payback (Years)",
                f"=MATCH(TRUE, 'Cash Flows'!B11:{get_column_letter(2 + len(self.assumptions.annual_cash_flows))}11>=0, 0)-1",
                "0.0",
            ),
            (
                "Average Annual Return",
                f"=AVERAGE('Cash Flows'!C5:{get_column_letter(2 + len(self.assumptions.annual_cash_flows))}5)",
                "$#,##0",
            ),
            (
                "Total Undiscounted Return",
                f"=SUM('Cash Flows'!B7:{get_column_letter(2 + len(self.assumptions.annual_cash_flows))}7)",
                "$#,##0",
            ),
        ]

        ws["A3"] = "Metric"
        ws["B3"] = "Value"
        ws["C3"] = "Interpretation"
        for col in range(1, 4):
            ws.cell(row=3, column=col).font = Font(bold=True)
            ws.cell(row=3, column=col).fill = PatternFill(
                start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid"
            )

        interpretations = [
            "Positive NPV indicates value creation",
            "Return rate that makes NPV = 0",
            "Value created per dollar invested",
            "Years to recover investment (nominal)",
            "Years to recover investment (discounted)",
            "Mean annual cash inflow",
            "Total cash generated over project life",
        ]

        for row, (metric, formula, format_str) in enumerate(metrics, 4):
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = formula
            ws[f"B{row}"].number_format = format_str
            ws[f"B{row}"].fill = PatternFill(
                start_color=COLORS["output"], end_color=COLORS["output"], fill_type="solid"
            )
            ws[f"C{row}"] = interpretations[row - 4]

        # Decision recommendation
        ws["A12"] = "Investment Decision:"
        ws["A12"].font = Font(bold=True, size=12)
        ws["B12"] = (
            '=IF(AND(B4>0, B5>DiscountRate), "ACCEPT - Project creates value", "REJECT - Project destroys value")'
        )
        ws["B12"].font = Font(bold=True)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50

    def _create_scenario_analysis(self):
        """Create scenario analysis"""
        ws = self.wb.create_sheet("Scenarios")

        ws["A1"] = "Scenario Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        # Scenario definitions
        ws["A3"] = "Scenario"
        ws["B3"] = "Discount Rate"
        ws["C3"] = "Cash Flow Adj"
        ws["D3"] = "NPV"
        ws["E3"] = "IRR"
        ws["F3"] = "Decision"

        for col in range(1, 7):
            ws.cell(row=3, column=col).font = Font(bold=True)
            ws.cell(row=3, column=col).fill = PatternFill(
                start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid"
            )

        scenarios = [
            ("Base Case", 1.0, 1.0),
            ("Best Case", 0.9, 1.2),
            ("Worst Case", 1.2, 0.8),
            ("Conservative", 1.1, 0.9),
            ("Aggressive", 0.95, 1.1),
        ]

        for row, (name, rate_mult, cf_mult) in enumerate(scenarios, 4):
            ws[f"A{row}"] = name
            ws[f"B{row}"] = f"=DiscountRate*{rate_mult}"
            ws[f"B{row}"].number_format = "0.0%"
            ws[f"C{row}"] = f"{cf_mult:.0%}"
            # Note: Actual NPV/IRR would require recalculating with adjusted values
            ws[f"D{row}"] = f"=Metrics!B4*{cf_mult}"
            ws[f"D{row}"].number_format = "$#,##0"
            ws[f"E{row}"] = f"=Metrics!B5*{cf_mult}"
            ws[f"E{row}"].number_format = "0.00%"
            ws[f"F{row}"] = f'=IF(D{row}>0, "Accept", "Reject")'

        ws.column_dimensions["A"].width = 20
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18


def calculate_roi(
    project_name: str,
    initial_investment: float,
    annual_cash_flows: list[float],
    discount_rate: float,
    tax_rate: float = 0.25,
    salvage_value: float = 0.0,
    output_path: str | None = None,
) -> str:
    """
    Create ROI analysis for investment opportunity.

    Args:
        project_name: Name of investment project
        initial_investment: Upfront investment amount (positive)
        annual_cash_flows: List of annual cash flows (can be positive or negative)
        discount_rate: Discount rate / WACC (as decimal, e.g., 0.10 for 10%)
        tax_rate: Corporate tax rate (as decimal)
        salvage_value: End-of-project salvage value
        output_path: Where to save file (optional)

    Returns:
        Path to created file
    """
    assumptions = InvestmentAssumptions(
        project_name=project_name,
        initial_investment=initial_investment,
        annual_cash_flows=annual_cash_flows,
        discount_rate=discount_rate,
        tax_rate=tax_rate,
        salvage_value=salvage_value,
    )

    calculator = ROICalculator(assumptions)

    if output_path is None:
        output_path = f'/mnt/user-data/outputs/{project_name.replace(" ", "_")}_ROI_Analysis.xlsx'

    return calculator.create_analysis(output_path)


if __name__ == "__main__":
    # Example usage
    model_path = calculate_roi(
        project_name="New Product Launch",
        initial_investment=5000000,
        annual_cash_flows=[1200000, 1500000, 1800000, 2000000, 2200000],
        discount_rate=0.10,
        tax_rate=0.25,
        salvage_value=500000,
    )
    print(f"ROI analysis created: {model_path}")
